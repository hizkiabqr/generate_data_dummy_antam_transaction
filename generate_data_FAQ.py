import pandas as pd
from datetime import datetime

# Data FAQ PT Antam
faq_data = [
    # Informasi Umum
    {
        'Kategori': 'Informasi Umum',
        'Pertanyaan': 'Apa itu PT Antam (Aneka Tambang Tbk)?',
        'Jawaban': 'PT Aneka Tambang Tbk (Antam) adalah perusahaan publik milik negara (BUMN) yang bergerak di bidang pertambangan dan pengolahan mineral strategis, terutama nikel, bauksit, dan emas. Antam merupakan anggota MIND ID (Mining Industry Indonesia), holding industri pertambangan nasional. Didirikan pada 5 Juli 1968.'
    },
    {
        'Kategori': 'Informasi Umum',
        'Pertanyaan': 'Apa saja unit bisnis dan segmen operasi utama Antam?',
        'Jawaban': 'Antam membagi bisnisnya ke dalam tiga segmen utama: 1) Emas dan Logam Mulia, 2) Nikel (termasuk ferronickel dan bijih nikel), 3) Bauksit dan Alumina'
    },
    {
        'Kategori': 'Informasi Umum',
        'Pertanyaan': 'Di mana lokasi tambang emas utama Antam?',
        'Jawaban': 'Tambang emas utama Antam berada di Pongkor, Jawa Barat, dan Cibaliung, Banten (yang kini memasuki masa pasca-tambang)'
    },
    {
        'Kategori': 'Informasi Umum',
        'Pertanyaan': 'Berapa pendapatan dan laba Antam terbaru?',
        'Jawaban': 'Pada tahun 2024, Antam membukukan pendapatan Rp 69,19 triliun dan laba bersih Rp 3,85 triliun, dengan total aset Rp 44,52 triliun'
    },
    
    # Produk Logam Mulia
    {
        'Kategori': 'Produk Logam Mulia',
        'Pertanyaan': 'Apa saja produk logam mulia yang ditawarkan Antam?',
        'Jawaban': 'Antam menawarkan: Emas Batangan Antam LM (0,5g-1000g), Emas Batik Series (10g & 20g), Emas Gift Series (0,5g & 1g), Emas Imlek dan Idul Fitri Series, Perak Batangan (99,95%), Dinar, Dirham, Platinum labware, dan produk custom'
    },
    {
        'Kategori': 'Produk Logam Mulia',
        'Pertanyaan': 'Apa spesifikasi dan keunggulan produk emas Antam?',
        'Jawaban': 'Emas Antam LM memiliki kemurnian 99,99% (24 karat), diakui internasional dengan sertifikasi LBMA. Setiap emas dilengkapi sertifikat keaslian, nomor seri unik, dan kemasan CertiCard dengan QR Code untuk verifikasi digital via aplikasi CertiEye'
    },
    {
        'Kategori': 'Produk Logam Mulia',
        'Pertanyaan': 'Apa itu CertiCard dan CertiEye pada produk Antam?',
        'Jawaban': 'CertiCard adalah kemasan pelindung emas batangan Antam yang tersegel dan tidak bisa dibuka tanpa merusak kemasan. Di bagian belakang terdapat QR Code yang dapat dipindai menggunakan aplikasi CertiEye untuk memverifikasi keaslian produk secara digital'
    },
    {
        'Kategori': 'Produk Logam Mulia',
        'Pertanyaan': 'Apakah produk emas Antam diakui secara internasional?',
        'Jawaban': 'Ya, emas batangan Antam LM telah memperoleh sertifikasi LBMA, sehingga diakui di pasar internasional dan dapat diperjualbelikan di luar negeri tanpa hambatan'
    },

    # Harga
    {
        'Kategori': 'Harga',
        'Pertanyaan': 'Berapa harga jual emas Antam hari ini?',
        'Jawaban': 'Per 6 Januari 2026, harga emas Antam LM 1 gram adalah Rp2.549.000 per gram (naik Rp34.000 dari hari sebelumnya). Harga buyback Rp2.405.000 per gram'
    },
    {
        'Kategori': 'Harga',
        'Pertanyaan': 'Berapa harga jual perak Antam hari ini?',
        'Jawaban': 'Harga perak Antam pada 6 Januari 2026 tercatat Rp47.265 per gram (naik Rp1.000 dari hari sebelumnya), sudah termasuk PPN 11%'
    },
    {
        'Kategori': 'Harga',
        'Pertanyaan': 'Apakah harga emas dan perak Antam selalu mengikuti harga internasional?',
        'Jawaban': 'Harga emas dan perak Antam umumnya mengikuti tren harga internasional, namun dipengaruhi oleh kurs Rupiah terhadap Dolar AS, biaya produksi, dan permintaan domestik'
    },
    {
        'Kategori': 'Harga',
        'Pertanyaan': 'Bagaimana tren harga emas Antam sepanjang tahun?',
        'Jawaban': 'Harga emas Antam cenderung naik sepanjang 2025 dan awal 2026, dipicu oleh ketegangan geopolitik global, ekspektasi penurunan suku bunga The Fed, dan pelemahan Dolar AS. Namun harga juga fluktuatif dan bisa turun pada periode tertentu'
    },

    # Tempat Pembelian
    {
        'Kategori': 'Tempat Pembelian',
        'Pertanyaan': 'Di mana saja tempat resmi membeli emas Antam asli?',
        'Jawaban': 'Tempat resmi: Butik Emas Logam Mulia (BELM) Antam di kota besar, Unit Bisnis Antam Pulogadung, Pegadaian dan Galeri 24, Bank Syariah, Butik Emas Mobile, dan Marketplace Resmi terverifikasi'
    },
    {
        'Kategori': 'Tempat Pembelian',
        'Pertanyaan': 'Bagaimana cara memastikan toko online atau marketplace resmi?',
        'Jawaban': 'Pastikan toko memiliki reputasi baik, mencantumkan sertifikat keaslian, kwitansi resmi, dan tidak menawarkan harga jauh di bawah pasaran. Hindari transaksi via WhatsApp atau rekening pribadi yang tidak resmi'
    },

    # Prosedur Pembelian
    {
        'Kategori': 'Prosedur Pembelian',
        'Pertanyaan': 'Bagaimana cara membeli emas Antam secara langsung di butik?',
        'Jawaban': 'Langkah: 1) Datang ke butik dengan KTP, 2) Ambil nomor antrean dan isi formulir, 3) Pilih produk dan gramasi, 4) Lakukan pembayaran (tunai/transfer), 5) Terima emas dan sertifikat keaslian'
    },
    {
        'Kategori': 'Prosedur Pembelian',
        'Pertanyaan': 'Bagaimana cara membeli emas Antam secara online di LogamMulia.com?',
        'Jawaban': 'Prosedur: 1) Registrasi akun di logammulia.com, 2) Pilih produk dan gramasi, 3) Pembayaran via virtual account bank, 4) Emas dikirim ke alamat atau ambil di butik, 5) Terima sertifikat dan kemasan utuh'
    },
    {
        'Kategori': 'Prosedur Pembelian',
        'Pertanyaan': 'Apa itu BRANKAS digital dan bagaimana cara membeli emas di sana?',
        'Jawaban': 'BRANKAS (Berencana Aman Kelola Emas) adalah layanan investasi emas digital dari Antam. Emas dicatat digital dan disimpan di fasilitas Antam. Prosedur: registrasi via app/website, isi data diri, pilih keanggotaan, beli emas digital, dapat dicetak fisik kapan saja'
    },
    {
        'Kategori': 'Prosedur Pembelian',
        'Pertanyaan': 'Apakah bisa membeli emas Antam secara cicilan?',
        'Jawaban': 'Pembelian cicilan dapat dilakukan melalui Pegadaian, Galeri 24, atau bank syariah. Di Butik Emas LM Antam pembelian umumnya tunai atau transfer penuh'
    },

    # Buyback
    {
        'Kategori': 'Buyback',
        'Pertanyaan': 'Bagaimana cara menjual kembali (buyback) emas Antam di butik resmi?',
        'Jawaban': 'Langkah: 1) Datang ke butik dengan KTP dan NPWP (jika ada), 2) Ambil nomor antrean dan isi formulir, 3) Serahkan emas, sertifikat, dan kemasan utuh, 4) Petugas periksa fisik dan sertifikat, 5) Dana ditransfer H+1 sampai H+3 hari kerja'
    },
    {
        'Kategori': 'Buyback',
        'Pertanyaan': 'Apa syarat buyback emas Antam?',
        'Jawaban': 'Syarat: kemasan CertiCard utuh dan tersegel (produk baru), sertifikat asli disertakan (produk retro), jika rusak/hilang ada potongan biaya, transaksi >Rp10 juta dikenakan PPh 22 sebesar 1,5% (NPWP) atau 3% (non-NPWP)'
    },
    {
        'Kategori': 'Buyback',
        'Pertanyaan': 'Apakah bisa melakukan buyback emas Antam secara online?',
        'Jawaban': 'Buyback emas digital dapat dilakukan melalui aplikasi BRANKAS atau platform tabungan emas di Pegadaian. Emas fisik dapat dicetak dan dijual kembali di butik atau via layanan buyback online Antam'
    },

    # Biaya dan Pajak
    {
        'Kategori': 'Biaya dan Pajak',
        'Pertanyaan': 'Apa saja biaya dan pajak saat membeli emas Antam?',
        'Jawaban': 'Biaya: PPh 22 sebesar 0,45% (NPWP) atau 0,9% (non-NPWP) sesuai PMK 34/PMK.10/2017, biaya cetak untuk produk tertentu, biaya administrasi BRANKAS mulai Rp100.000/tahun. PPN hanya untuk perhiasan, tidak untuk emas batangan'
    },
    {
        'Kategori': 'Biaya dan Pajak',
        'Pertanyaan': 'Apa saja biaya dan pajak saat jual kembali (buyback) emas Antam?',
        'Jawaban': 'Biaya: PPh 22 Buyback 1,5% (NPWP) atau 3% (non-NPWP) untuk transaksi >Rp10 juta, potongan kemasan/sertifikat jika rusak atau hilang sesuai ketentuan butik'
    },
    {
        'Kategori': 'Biaya dan Pajak',
        'Pertanyaan': 'Apakah perak batangan Antam dikenakan PPN?',
        'Jawaban': 'Ya, perak batangan Antam dikenakan PPN 11% pada harga jual, sesuai ketentuan pajak terbaru'
    },

    # Keaslian
    {
        'Kategori': 'Keaslian',
        'Pertanyaan': 'Bagaimana cara mengecek keaslian emas Antam CertiCard?',
        'Jawaban': 'Cara: 1) Scan QR Code dengan aplikasi CertiEye (logo Antam LM dan "AUTHENTICATED" jika asli), 2) Periksa fisik emas (cap Fine Gold 999.9, nomor seri, logo), 3) Cek sertifikat dengan sinar UV (produk retro), 4) Verifikasi nomor seri di logammulia.com'
    },
    {
        'Kategori': 'Keaslian',
        'Pertanyaan': 'Apa yang harus dilakukan jika hasil verifikasi CertiEye tidak muncul?',
        'Jawaban': 'Segera hubungi hotline resmi Antam (Call Center 0804-1-888-888, WhatsApp ALMIRA 0811-1002-002) untuk konfirmasi dan tindak lanjut'
    },
    {
        'Kategori': 'Keaslian',
        'Pertanyaan': 'Bagaimana cara menghindari penipuan dan produk palsu?',
        'Jawaban': 'Tips: selalu beli di tempat resmi, jangan tergiur harga murah, hindari transaksi via WhatsApp atau rekening pribadi tidak resmi, cek keaslian sertifikat dan kemasan, simpan bukti pembelian'
    },

    # Penyimpanan
    {
        'Kategori': 'Penyimpanan',
        'Pertanyaan': 'Bagaimana cara menyimpan emas Antam agar aman dan tidak rusak?',
        'Jawaban': 'Cara: 1) Simpan di brankas rumah (kotak besi berlapis, gembok kombinasi, tempat tersembunyi), 2) Safe Deposit Box (SDB) di bank, 3) Jasa penyimpanan BRANKAS LM, 4) Jangan beritahu kepemilikan ke orang lain, 5) Asuransikan emas untuk jumlah besar'
    },
    {
        'Kategori': 'Penyimpanan',
        'Pertanyaan': 'Bagaimana cara menyimpan sertifikat emas Antam?',
        'Jawaban': 'Simpan sertifikat di tempat terpisah dari emas fisik, buat salinan digital sebagai cadangan, dan hindari kerusakan fisik pada dokumen'
    },

    # Strategi Investasi
    {
        'Kategori': 'Strategi Investasi',
        'Pertanyaan': 'Apa itu strategi Dollar Cost Averaging (DCA) untuk investasi emas?',
        'Jawaban': 'DCA adalah strategi investasi dengan membeli emas secara rutin dalam jumlah uang yang sama, tanpa memperhatikan harga pasar. DCA membantu merata-ratakan harga beli dan mengurangi risiko volatilitas harga emas. Cocok untuk pemula dan jangka panjang'
    },
    {
        'Kategori': 'Strategi Investasi',
        'Pertanyaan': 'Apakah investasi emas Antam cocok untuk jangka pendek atau panjang?',
        'Jawaban': 'Investasi emas Antam paling optimal untuk jangka menengah hingga panjang (minimal 3-5 tahun, idealnya 10 tahun ke atas). Kenaikan harga emas cenderung lambat dan spread harga beli-jual cukup tinggi, kurang cocok untuk spekulasi jangka pendek'
    },
    {
        'Kategori': 'Strategi Investasi',
        'Pertanyaan': 'Bagaimana cara diversifikasi investasi emas?',
        'Jawaban': 'Jangan menaruh seluruh modal di emas, kombinasikan dengan instrumen lain seperti reksa dana, deposito, atau saham. Diversifikasi membantu mengurangi risiko dan meningkatkan potensi keuntungan portofolio'
    },
    {
        'Kategori': 'Strategi Investasi',
        'Pertanyaan': 'Apa saja ukuran gramasi emas Antam yang direkomendasikan untuk pemula?',
        'Jawaban': 'Rekomendasi: 0,5 gram (coba-coba), 1 gram (favorit pemula, mudah dijual), 2 gram (lebih hemat biaya cetak), 5 gram (jangka menengah), 10 gram (pemula ambisius dengan budget lebih besar)'
    },

    # Waktu Terbaik Membeli
    {
        'Kategori': 'Waktu Terbaik Membeli',
        'Pertanyaan': 'Kapan waktu terbaik membeli emas Antam menurut tren musiman?',
        'Jawaban': 'Waktu terbaik: saat harga turun dari puncak, awal tahun (Januari-Maret), April-Juni (harga stabil/turun), sebelum puncak permintaan Imlek/Lebaran, saat Dolar AS menguat terhadap Rupiah'
    },
    {
        'Kategori': 'Waktu Terbaik Membeli',
        'Pertanyaan': 'Apa saja faktor utama yang mempengaruhi harga emas?',
        'Jawaban': 'Faktor: nilai tukar Dolar AS, kondisi ekonomi global (inflasi, suku bunga, pertumbuhan), geopolitik (ketegangan, perang, krisis), permintaan domestik (musim perayaan, tradisi), aksi bank sentral (pembelian cadangan emas)'
    },
    {
        'Kategori': 'Waktu Terbaik Membeli',
        'Pertanyaan': 'Kalau musim Imlek, Lebaran, atau pernikahan, baiknya beli atau jual emas?',
        'Jawaban': 'Imlek & Lebaran: permintaan naik, harga melonjak. Beli sebelum puncak permintaan, jual saat permintaan tinggi untuk hasil maksimal. Pernikahan & musim panen: jual saat permintaan tinggi. Awal tahun & April-Juni: waktu baik untuk akumulasi investasi'
    },

    # Pemula
    {
        'Kategori': 'Pemula',
        'Pertanyaan': 'Bagaimana cara memilih ukuran emas Antam sesuai budget dan tujuan?',
        'Jawaban': 'Budget terbatas: 0,5-2 gram (fleksibilitas). Jangka menengah: 5-10 gram (efisiensi biaya cetak). Jangka panjang: 25-1000 gram (potensi maksimal, perdagangan internasional). Hadiah/koleksi: Gift Series atau Batik Series'
    },
    {
        'Kategori': 'Pemula',
        'Pertanyaan': 'Apa saja tips membeli emas Antam untuk pemula?',
        'Jawaban': 'Tips: cek harga harian di situs resmi sebelum beli, beli di tempat resmi, simpan sertifikat dan kemasan dengan baik, diversifikasi investasi, rencanakan investasi jangka panjang (3-10 tahun)'
    },
    {
        'Kategori': 'Pemula',
        'Pertanyaan': 'Apa perbedaan emas batangan Antam dengan perhiasan emas untuk investasi?',
        'Jawaban': 'Emas batangan Antam: kemurnian lebih tinggi (99,99%), harga jual kembali mengikuti pasar tanpa potongan biaya produksi, sertifikat berstandar internasional. Perhiasan: cocok untuk gaya hidup, harga jual kembali lebih rendah karena potongan ongkos produksi dan kadar emas lebih rendah'
    },

    # Kontak
    {
        'Kategori': 'Kontak',
        'Pertanyaan': 'Bagaimana cara menghubungi layanan pelanggan Antam?',
        'Jawaban': 'Call Center: 0804-1-888-888, WhatsApp ALMIRA: 0811-1002-002 (chat only), Email: infolm@antam.com, Kantor Pusat: Gedung Aneka Tambang Tower A, Jl. Letjen TB Simatupang No.1, Jakarta 12530'
    },
    {
        'Kategori': 'Kontak',
        'Pertanyaan': 'Apa yang harus dilakukan jika menerima produk palsu atau mengalami masalah transaksi?',
        'Jawaban': 'Segera hubungi Call Center Antam (0804-1-888-888), WhatsApp ALMIRA (0811-1002-002), atau email infolm@antam.com. Sertakan bukti pembelian, foto produk dan sertifikat, serta kronologi masalah. Antam akan melakukan verifikasi dan tindak lanjut'
    },

    # Wilayah Operasional
    {
        'Kategori': 'Wilayah Operasional',
        'Pertanyaan': 'Di mana saja lokasi tambang nikel PT Antam?',
        'Jawaban': 'Tambang nikel PT Antam berada di Pomalaa (Sulawesi Tenggara) dan Halmahera Timur (Maluku Utara). Pomalaa merupakan salah satu penghasil nikel terbesar di Indonesia dengan fasilitas pengolahan feronikel'
    },
    {
        'Kategori': 'Wilayah Operasional',
        'Pertanyaan': 'Di mana lokasi tambang bauksit PT Antam?',
        'Jawaban': 'Tambang bauksit PT Antam terletak di Tayan, Kalimantan Barat. Bauksit dari Tayan diolah menjadi alumina, bahan baku utama pembuatan aluminium, dan merupakan salah satu pusat produksi bauksit terbesar di Indonesia'
    },
    {
        'Kategori': 'Wilayah Operasional',
        'Pertanyaan': 'Apa kontribusi PT Antam terhadap pembangunan daerah?',
        'Jawaban': 'Kontribusi: menciptakan ribuan lapangan kerja, pembangunan infrastruktur (jalan, air bersih, listrik), program CSR fokus pada pendidikan, kesehatan, dan pemberdayaan ekonomi masyarakat, kerjasama dengan UKM lokal'
    },

    # Manajemen
    {
        'Kategori': 'Manajemen',
        'Pertanyaan': 'Siapa Direktur Utama PT Antam saat ini?',
        'Jawaban': 'Per 12 Juni 2025, Achmad Ardianto diangkat menjadi Direktur Utama PT Antam menggantikan direktur utama sebelumnya'
    },
    {
        'Kategori': 'Manajemen',
        'Pertanyaan': 'Siapa saja anggota Dewan Direksi PT Antam terbaru?',
        'Jawaban': 'Dewan Direksi: Achmad Ardianto (Dirut), Hartono (Dirops & Produksi), I Dewa Wirantaya (Pengembangan Usaha), Handi Sutanto (Komersial), Arianto Sabto Nugroho Rudjito (Keuangan & Risiko), Ratih Amri (SDM)'
    },
    {
        'Kategori': 'Manajemen',
        'Pertanyaan': 'Siapa Komisaris Utama PT Antam saat ini?',
        'Jawaban': 'Rauf Purnama menjabat sebagai Komisaris Utama merangkap Komisaris Independen PT Antam per 12 Juni 2025'
    }
]

# Buat DataFrame
df = pd.DataFrame(faq_data)

# Tambahkan nomor urut
df.insert(0, 'No', range(1, len(df) + 1))

# Simpan ke Excel dengan formatting
filename = f'FAQ_PT_Antam_{datetime.now().strftime("%Y%m%d")}.xlsx'

# Buat Excel writer object
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='FAQ PT Antam', index=False)
    
    # Ambil worksheet untuk formatting
    workbook = writer.book
    worksheet = writer.sheets['FAQ PT Antam']
    
    # Set lebar kolom
    worksheet.column_dimensions['A'].width = 5
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 50
    worksheet.column_dimensions['D'].width = 80
    
    # Format header
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    header_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format isi tabel
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
    
    # Freeze panes (header tetap terlihat saat scroll)
    worksheet.freeze_panes = 'A2'

print(f'✅ File Excel berhasil dibuat: {filename}')
print(f'📊 Total FAQ: {len(df)} pertanyaan')
print(f'📁 Kategori: {df["Kategori"].nunique()} kategori')
print('\n🎯 Kategori yang tersedia:')
for idx, kategori in enumerate(df['Kategori'].unique(), 1):
    count = len(df[df['Kategori'] == kategori])
    print(f'   {idx}. {kategori}: {count} FAQ')